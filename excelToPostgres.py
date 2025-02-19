import openpyxl

# Путь к файлу Excel
path = "/Users/oxydear/Documents/Ivan's Mac/IDEs/PyCharm/problemsSolve/basesScript/atomar/Текущий рейтинг студентов ФИТУ .xlsx"
wb_obj = openpyxl.load_workbook(path)

# Активный лист
sheet_obj = wb_obj.active

# Максимальное количество строк и столбцов
row = sheet_obj.max_row
column = sheet_obj.max_column
arr = []

for cl in range(1, column + 1):
    cell_obj = sheet_obj.cell(row=1, column=cl)
    value = cell_obj.value
    arr.append(value)

with open("../insertAtom.txt", "w") as file:
    for rw in range(2, row + 1):
        insert = ""
        for cl in range(3, column + 1):
            cell_obj = sheet_obj.cell(row=rw, column=cl)
            group = sheet_obj.cell(row=rw, column=1).value
            stud = sheet_obj.cell(row=rw, column=2).value
            value = cell_obj.value

            insert += f"INSERT INTO students VALUES ({group}, '{stud}', '{arr[cl-1]}', '{value}');\n"
        file.write(insert)


